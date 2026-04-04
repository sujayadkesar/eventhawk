"""
PowerShell forensic extraction — output file writers.

Produces five files in the user-chosen output directory:
  ps_commands.txt           — chronological command/session timeline
  scriptblock_<GUID>.txt    — one per unique ScriptBlockId (reassembled)
  ps_extraction_summary.txt — statistics and ATT&CK summary
  ps_extraction.json        — machine-readable SIEM export
  ps_timeline.xlsx          — flat Excel timeline with clickable script-block links
"""

from __future__ import annotations

import json
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable

from .constants import NOTABLE_PROVIDERS, ZERO_GUID
from .models import ContentAnalysisResult, PSSession, ScriptBlockAccumulator

_SEP = "=" * 80
_SEP_THIN = "-" * 80
_LARGE_BLOCK_THRESHOLD = 10 * 1024 * 1024  # 10 MB


def _fmt_ts(ts: str) -> str:
    """Reformat ISO-8601 timestamp to human-readable 'YYYY-MM-DD HH:MM:SS.ffffff'."""
    if not ts:
        return ""
    return ts.replace("T", " ").rstrip("Z")


def _is_ghost_session(sess: PSSession) -> bool:
    """
    Return True for synthetic sessions with no parseable metadata.

    These arise from EID 400 records whose HostId field was empty or corrupt
    (session_builder gives them a ``_pid{pid}_{timestamp}`` fallback key).
    When they also have no host_name, no host_application, and only the single
    EID 400 event correlated to them they carry no forensic value and are
    excluded from ps_commands.txt and the JSON/CSV exports.  They are still
    counted in ps_extraction_summary.txt.
    """
    return (
        sess.host_id.startswith("_pid")
        and not sess.host_name
        and not sess.host_application
        and len(sess.session_events) <= 1
    )


# ── ps_commands.txt ───────────────────────────────────────────────────────────

def write_commands_file(
    sessions: list[PSSession],
    buckets: dict[int, list[dict]],
    sb_index: dict[str, ScriptBlockAccumulator],
    analysis_results: dict[str, ContentAnalysisResult],
    output_path: Path,
    source_files: list[str],
) -> None:
    """
    Write ps_commands.txt — full session timeline with all events interleaved.
    """
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    source_names = ", ".join(os.path.basename(f) for f in source_files) or "(unknown)"

    lines: list[str] = [
        _SEP,
        "EventHawk — PowerShell Command Execution Log",
        f"Generated   : {now}",
        f"Source EVTX : {source_names}",
        _SEP,
        "",
    ]

    # Build reverse lookup: original_sbid → index key (handles ZERO_GUID synthetic keys).
    # For non-ZERO_GUID blocks, key == sbid. For ZERO_GUID blocks the key is
    # "{zero_guid}_{record_id}" — we cannot guess it from sbid alone, so we
    # resolve via the event's record_id (passed as a second argument below).
    def _resolve_key(sbid: str, record_id: int = 0) -> str:
        """Return the sb_index key for this sbid. For ZERO_GUID uses record_id."""
        if sbid == ZERO_GUID:
            return f"{sbid}_{record_id}"
        return sbid

    def _sb_flags(sbid: str, record_id: int = 0) -> str:
        key = _resolve_key(sbid, record_id)
        acc = sb_index.get(key)
        result = analysis_results.get(key)
        flags: list[str] = []
        if acc and acc.was_safety_net_triggered:
            flags.append("SAFETY_NET")
        if result:
            flags.extend(result.indicator_flags())
        return " ".join(f"[{f}]" for f in flags) if flags else ""

    def _sb_attck(sbid: str, record_id: int = 0) -> str:
        key = _resolve_key(sbid, record_id)
        result = analysis_results.get(key)
        if result and result.att_ck_techniques:
            return ", ".join(result.att_ck_techniques)
        return "T1059.001"

    def _sb_preview(sbid: str, record_id: int = 0) -> str:
        """Return a single-line content preview (first 300 chars) of a script block."""
        key = _resolve_key(sbid, record_id)
        acc = sb_index.get(key)
        if not acc:
            return ""
        text = acc.assemble().strip()
        if not text:
            return ""
        # Collapse internal whitespace to one space so the preview stays on one line
        flat = " ".join(text.split())
        total = len(text)
        if len(flat) > 300:
            return flat[:300] + f"... [{total:,} chars — see scriptblock_{key[:8]}….txt]"
        return flat

    for sess in sessions:
        if sess.host_id == "_unsessioned":
            continue  # written separately at end
        if _is_ghost_session(sess):
            continue  # corrupt EID 400 with no metadata — omit from timeline
        _write_session_block(lines, sess, _sb_flags, _sb_attck, _sb_preview)

    # Unsessioned events
    unsessioned_evs: list[dict] = []
    for sess in sessions:
        if sess.host_id == "_unsessioned":
            unsessioned_evs = sess.session_events
            break

    if unsessioned_evs:
        lines += [
            "",
            _SEP_THIN,
            "--- UNSESSIONED EVENTS (no matching EID 400 engine-start) ---",
            _SEP_THIN,
            "",
        ]
        for ev in unsessioned_evs:
            lines.append(_format_event_line(ev, _sb_flags, _sb_attck, _sb_preview))
        lines.append("")

    output_path.write_text("\n".join(lines), encoding="utf-8", errors="replace")


def _write_session_block(
    lines: list[str],
    sess: PSSession,
    sb_flags_fn: Callable[[str, int], str],
    sb_attck_fn: Callable[[str, int], str],
    sb_preview_fn: Callable[[str, int], str] | None = None,
) -> None:
    duration = sess.duration_str()
    dur_str = f" | Duration={duration}" if duration else ""
    end_str = _fmt_ts(sess.stop_ts) if sess.stop_ts else "(no stop event)"

    lines += [
        "",
        f"--- SESSION: HostId={sess.host_id} | PID={sess.pid}"
        f" | Start={_fmt_ts(sess.start_ts)} | End={end_str}{dur_str} ---",
        f"  Host        : {sess.host_name} v{sess.host_version}",
        f"  HostApplication: {sess.host_application}",
    ]

    if sess.user_sid or sess.user_name:
        user_str = sess.user_name if sess.user_name else "(unknown)"
        sid_str  = sess.user_sid  if sess.user_sid  else "(unknown)"
        lines.append(f"  User        : {user_str}  SID: {sid_str}")

    if sess.encoded_command:
        # Truncate long decoded commands to 500 chars for readability
        decoded_preview = sess.encoded_command[:500]
        if len(sess.encoded_command) > 500:
            decoded_preview += f"... [{len(sess.encoded_command)} chars total]"
        lines.append(f"  [ENCODED_COMMAND_DETECTED] Decoded: {decoded_preview}")

    if sess.providers:
        notable = [p for p in sess.providers if p in NOTABLE_PROVIDERS]
        prov_str = ", ".join(sess.providers)
        if notable:
            prov_str += f"  [NOTABLE: {', '.join(notable)}]"
        lines.append(f"  Providers   : {prov_str}")

    lines.append("")

    for ev in sess.session_events:
        lines.append(_format_event_line(ev, sb_flags_fn, sb_attck_fn, sb_preview_fn))

    lines.append("")
    lines.append("--- END SESSION ---")


def _format_event_line(
    ev: dict,
    sb_flags_fn: Callable[[str, int], str],
    sb_attck_fn: Callable[[str, int], str],
    sb_preview_fn: Callable[[str, int], str] | None = None,
) -> str:
    eid = ev.get("event_id", "?")
    ts = _fmt_ts(ev.get("timestamp", ""))
    pid = ev.get("pid", "")

    if eid == 400:
        return (
            f"[{ts}] EID:400  ENGINE_START     PID:{pid}"
            f"  State: {ev.get('prev_engine_state','?')} -> {ev.get('new_engine_state','?')}"
        )
    if eid == 403:
        return (
            f"[{ts}] EID:403  ENGINE_STOP      PID:{pid}"
            f"  State: {ev.get('prev_engine_state','?')} -> {ev.get('new_engine_state','?')}"
        )
    if eid == 600:
        return (
            f"[{ts}] EID:600  PROVIDER_START   PID:{pid}"
            f"  Provider: {ev.get('provider_name','?')}"
        )
    if eid == 4103:
        cmd = ev.get("command_name", "")
        cmd_type = ev.get("command_type", "")
        runspace = ev.get("runspace_id", "")
        payload = ev.get("payload", "") or ""
        script = ev.get("script_name", "") or "(interactive)"
        # Truncate payload preview
        payload_preview = payload[:200] + ("..." if len(payload) > 200 else "")
        result = [
            f"[{ts}] EID:4103 COMMAND_EXEC     PID:{pid}  Runspace:{runspace}",
            f"  Command : {cmd}",
            f"  Type    : {cmd_type}",
            f"  Script  : {script}",
        ]
        if payload_preview:
            result.append(f"  Payload : {payload_preview}")
        return "\n".join(result)
    if eid == 4104:
        sbid = ev.get("script_block_id", "")
        record_id = int(ev.get("event_record_id", 0))
        msg_num = ev.get("message_number", 1)
        msg_tot = ev.get("message_total", 1)
        path = ev.get("path", "") or "(in-memory)"
        # Pass record_id so the closure can reconstruct synthetic ZERO_GUID keys
        flags = sb_flags_fn(sbid, record_id)
        attck = sb_attck_fn(sbid, record_id)
        result = [
            f"[{ts}] EID:4104 SCRIPTBLOCK      PID:{pid}  Block:{sbid} [{msg_num}/{msg_tot}]",
            f"  Path    : {path}",
        ]
        if flags:
            result.append(f"  Flags   : {flags}")
        result.append(f"  ATT&CK  : {attck}")
        if sb_preview_fn:
            preview = sb_preview_fn(sbid, record_id)
            if preview:
                result.append(f"  Content : {preview}")
        return "\n".join(result)
    if eid == 800:
        cmd = ev.get("command_name", "")
        cmd_line = ev.get("command_line", "")[:200]
        return (
            f"[{ts}] EID:800  PIPELINE_DETAIL  PID:{pid}"
            f"  Cmd:{cmd}  Line:{cmd_line}"
        )
    return f"[{ts}] EID:{eid}  PID:{pid}"


# ── scriptblock_<GUID>.txt ────────────────────────────────────────────────────

def write_scriptblock_files(
    sb_index: dict[str, ScriptBlockAccumulator],
    analysis_results: dict[str, ContentAnalysisResult],
    output_dir: Path,
    progress_cb: Callable[[str, float], None] | None = None,
    errors: list[str] | None = None,
) -> None:
    """
    Write one scriptblock_<GUID>.txt file per ScriptBlockAccumulator.
    Emits progress every 10% via progress_cb(step, pct) where pct is 0.78->0.92.
    """
    total = len(sb_index)
    if total == 0:
        return

    chunk = max(1, total // 10)

    for i, (key, acc) in enumerate(sb_index.items()):
        if progress_cb and i % chunk == 0:
            pct = 0.78 + (i / total) * 0.14
            progress_cb(f"Writing script block files ({i}/{total})...", pct)

        result = analysis_results.get(key)
        # Sanitise filename: strip braces, replace colons/slashes
        safe_key = key.replace("{", "").replace("}", "").replace("/", "_").replace("\\", "_")
        # Truncate to avoid filesystem limits (255 char filename limit on most OSes)
        if len(safe_key) > 200:
            safe_key = safe_key[:200]
        outpath = output_dir / f"scriptblock_{safe_key}.txt"

        try:
            content = _format_scriptblock_file(acc, result)
            outpath.write_text(content, encoding="utf-8", errors="replace")
        except OSError as exc:
            msg = f"Failed to write {outpath}: {exc}"
            if errors is not None:
                errors.append(msg)


def _format_scriptblock_file(
    acc: ScriptBlockAccumulator,
    result: ContentAnalysisResult | None,
) -> str:
    assembled = acc.assemble()
    size_bytes = len(assembled.encode("utf-8", errors="replace"))
    large_note = f"  [LARGE_BLOCK: {size_bytes:,} bytes]\n" if size_bytes > _LARGE_BLOCK_THRESHOLD else ""

    total = acc.expected_total or 1
    present = len(acc.fragments)
    missing = acc.missing_count()
    if acc.is_complete:
        frag_status = f"{present} / {total} (COMPLETE)"
    else:
        frag_status = f"{present} / {total} (INCOMPLETE — {missing} fragment(s) missing)"

    level_str = "WARNING (Auto-Safety-Net Triggered)" if acc.was_safety_net_triggered else "VERBOSE"
    path_str = acc.path if acc.path else "(in-memory / interactive)"

    attck_str = ""
    indicators_str = ""
    if result:
        attck_str = ", ".join(result.att_ck_techniques) if result.att_ck_techniques else "T1059.001"
        flags = result.indicator_flags()
        indicators_str = ", ".join(flags) if flags else "(none)"
        if result.detected_patterns:
            pattern_lines = "\n".join(f"  - {p}" for p in sorted(set(result.detected_patterns)))
        else:
            pattern_lines = "  (none)"
    else:
        attck_str = "T1059.001"
        indicators_str = "(none)"
        pattern_lines = "  (none)"

    # Detect Protected Event Logging
    pel_note = ""
    if "EncryptedMessage" in assembled or "EncryptedContent" in assembled:
        pel_note = "\n[PROTECTED_EVENT_LOGGING] — ScriptBlockText contains encrypted content.\n"

    header = "\n".join([
        _SEP,
        "EventHawk — Reconstructed Script Block",
        _SEP,
        f"ScriptBlockId  : {acc.script_block_id}",
        f"Path           : {path_str}",
        f"Computer       : {acc.computer}",
        f"FragmentCount  : {frag_status}",
        f"FirstSeen      : {_fmt_ts(acc.first_timestamp)}",
        f"LastSeen       : {_fmt_ts(acc.last_timestamp)}",
        f"LoggingLevel   : {level_str}",
        f"ATT&CK         : {attck_str}",
        f"Indicators     : {indicators_str}",
        "",
        "Detected Patterns:",
        pattern_lines,
        "",
    ])

    return (
        header
        + large_note
        + pel_note
        + _SEP + "\n"
        + "--- SCRIPT CONTENT BEGIN ---\n"
        + _SEP + "\n\n"
        + assembled
        + "\n\n"
        + _SEP + "\n"
        + "--- SCRIPT CONTENT END ---\n"
        + _SEP + "\n"
    )


# ── ps_extraction_summary.txt ─────────────────────────────────────────────────

def write_summary(
    buckets: dict[int, list[dict]],
    sb_index: dict[str, ScriptBlockAccumulator],
    sessions: list[PSSession],
    analysis_results: dict[str, ContentAnalysisResult],
    output_path: Path,
    source_files: list[str],
    total_scanned: int,
    errors: list[str],
) -> dict:
    """
    Write ps_extraction_summary.txt and return a summary dict for the GUI.
    """
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    source_names = ", ".join(os.path.basename(f) for f in source_files) or "(unknown)"

    counts = {eid: len(evs) for eid, evs in buckets.items()}
    total_ps = sum(counts.values())

    # Script block stats
    complete_blocks = [acc for acc in sb_index.values() if acc.is_complete]
    partial_blocks = [acc for acc in sb_index.values() if not acc.is_complete]
    single_frag = [acc for acc in sb_index.values() if acc.is_single_fragment]
    multi_frag_complete = [acc for acc in complete_blocks if not acc.is_single_fragment]
    safety_net_blocks = [acc for acc in sb_index.values() if acc.was_safety_net_triggered]
    # Detect PS Core (pwsh.exe) blocks by checking the source channel stored on
    # each fragment (ScriptFragment.channel). Fragments from PowerShellCore/Operational
    # have channel == "PowerShellCore/Operational". activity_id is a GUID and cannot
    # be used for channel detection.
    ps_core_blocks = [
        acc for acc in sb_index.values()
        if any(
            f.channel == "PowerShellCore/Operational"
            for f in acc.fragments.values()
        )
    ]

    # Session stats
    real_sessions = [s for s in sessions if s.host_id != "_unsessioned"]
    ghost_sessions = [s for s in real_sessions if _is_ghost_session(s)]
    real_sessions_with_data = [s for s in real_sessions if not _is_ghost_session(s)]
    sessions_with_stop = [s for s in real_sessions_with_data if s.stop_ts]
    sessions_without_stop = [s for s in real_sessions_with_data if not s.stop_ts]

    # Durations
    durations: list[float] = []
    for sess in sessions_with_stop:
        d = sess.duration_str()
        if d:
            try:
                parts = d.split(":")
                total_secs = int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
                durations.append(total_secs)
            except Exception:
                pass

    shortest_str = ""
    longest_str = ""
    if durations:
        def _fmt_secs(s: float) -> str:
            s = int(s)
            h, rem = divmod(s, 3600)
            m, sec = divmod(rem, 60)
            return f"{h}:{m:02d}:{sec:02d}"
        shortest_str = _fmt_secs(min(durations))
        longest_str = _fmt_secs(max(durations))

    # Provider stats across all sessions
    from collections import Counter
    provider_counts: Counter = Counter()
    for sess in real_sessions_with_data:
        for p in sess.providers:
            provider_counts[p] += 1

    # ATT&CK technique summary
    attck_counts: Counter = Counter()
    for result in analysis_results.values():
        for tid in result.att_ck_techniques:
            attck_counts[tid] += 1

    # Engine timeline (EID 400/403 pairs) — skip ghost sessions
    timeline_lines: list[str] = []
    for sess in real_sessions_with_data:
        dur = sess.duration_str()
        dur_note = f"  (duration: {dur})" if dur else "  (no stop event)"
        timeline_lines.append(
            f"  [{_fmt_ts(sess.start_ts)}] ENGINE_START PID:{sess.pid}  Host:{sess.host_name}"
        )
        if sess.stop_ts:
            timeline_lines.append(
                f"  [{_fmt_ts(sess.stop_ts)}] ENGINE_STOP  PID:{sess.pid}{dur_note}"
            )

    # Provider notable notes
    provider_lines: list[str] = []
    for prov, count in sorted(provider_counts.items(), key=lambda x: -x[1]):
        note = ""
        if prov == "Certificate":
            note = "  [NOTABLE: Certificate provider suggests cert enumeration or key theft]"
        elif prov == "WSMan":
            note = "  [NOTABLE: WinRM/remoting provider — lateral movement indicator]"
        elif prov == "ActiveDirectory":
            note = "  [NOTABLE: AD provider — reconnaissance or admin activity]"
        provider_lines.append(f"  {prov}: {count} session(s){note}")

    # ATT&CK table
    _attck_names = {
        "T1059.001": "PowerShell",
        "T1027":     "Obfuscated Files or Info",
        "T1105":     "Ingress Tool Transfer",
        "T1562.001": "Disable Security Tools (AMSI)",
        "T1055":     "Process Injection",
        "T1003":     "OS Credential Dumping",
        "T1620":     "Reflective Code Loading",
        "T1047":     "WMI",
        "T1053.005": "Scheduled Task/Job",
        "T1021.006": "WinRM (Lateral Movement)",
        "T1559.001": "COM Object",
    }
    attck_lines: list[str] = []
    for tid, cnt in sorted(attck_counts.items(), key=lambda x: -x[1]):
        name = _attck_names.get(tid, tid)
        attck_lines.append(f"  {tid}  {name:<40}: {cnt:>6} block(s)")

    # Parse errors
    error_lines: list[str] = []
    for err in errors[:50]:  # cap at 50 in file; full list in summary dict
        error_lines.append(f"  {err}")
    if len(errors) > 50:
        error_lines.append(f"  ... and {len(errors) - 50} more errors (see log)")

    lines = [
        _SEP,
        "EventHawk — PowerShell Extraction Summary",
        _SEP,
        f"Run At          : {now}",
        f"Source EVTX     : {source_names}",
        f"Source Events   : {total_scanned:,} total / {total_ps:,} PowerShell-related",
        "",
        "EVENT COUNTS BY TYPE:",
        f"  EID 4104 (Script Block Logging)    : {counts.get(4104, 0):>8,}",
        f"  EID 4103 (Module/Param Binding)    : {counts.get(4103, 0):>8,}",
        f"  EID 400  (Engine Start)            : {counts.get(400,  0):>8,}",
        f"  EID 403  (Engine Stop)             : {counts.get(403,  0):>8,}",
        f"  EID 600  (Provider Start)          : {counts.get(600,  0):>8,}",
        f"  EID 800  (Pipeline Details)        : {counts.get(800,  0):>8,}",
        "",
        "SCRIPT BLOCK ANALYSIS:",
        f"  Unique ScriptBlockIds              : {len(sb_index):>8,}",
        f"  Single-fragment blocks             : {len(single_frag):>8,}",
        f"  Multi-fragment blocks (complete)   : {len(multi_frag_complete):>8,}",
        f"  Partial blocks (missing fragments) : {len(partial_blocks):>8,}",
        f"  Safety-net auto-logged blocks      : {len(safety_net_blocks):>8,}",
        f"  PowerShell Core (pwsh.exe) blocks  : {len(ps_core_blocks):>8,}",
        "",
        "SESSION ANALYSIS:",
        f"  Total PS sessions (by EID 400)     : {len(real_sessions):>8,}",
        f"  Sessions with metadata (reported)  : {len(real_sessions_with_data):>8,}",
        f"  Ghost sessions (corrupt EID 400)   : {len(ghost_sessions):>8,}"
        + ("  [empty HostId — omitted from timeline/JSON]" if ghost_sessions else ""),
        f"  Sessions with matching EID 403     : {len(sessions_with_stop):>8,}",
        f"  Sessions without stop event        : {len(sessions_without_stop):>8,}"
        + ("  [possible forced termination]" if sessions_without_stop else ""),
    ]

    if shortest_str:
        lines.append(f"  Shortest session                   : {shortest_str}")
    if longest_str:
        lines.append(f"  Longest session                    : {longest_str}")

    if timeline_lines:
        lines += ["", "ENGINE TIMELINE:"] + timeline_lines

    if provider_lines:
        lines += ["", "PROVIDERS LOADED (by session):"] + provider_lines

    if attck_lines:
        lines += ["", "ATT&CK TECHNIQUE SUMMARY:"] + attck_lines

    lines += ["", f"PARSE ERRORS: {len(errors)}"]
    if error_lines:
        lines += error_lines
    else:
        lines.append("  (none)")

    lines += [
        "",
        "OUTPUT FILES:",
        f"  ps_commands.txt          (session timeline)",
        f"  scriptblock_<GUID>.txt x {len(sb_index)}",
        f"  ps_extraction_summary.txt",
        f"  ps_extraction.json       (machine-readable export)",
        f"  ps_timeline.xlsx         (Excel timeline with hyperlinks)",
        "",
        _SEP,
    ]

    output_path.write_text("\n".join(lines), encoding="utf-8", errors="replace")

    return {
        "total_scanned":    total_scanned,
        "total_ps_events":  total_ps,
        "script_blocks":    len(sb_index),
        "sessions":         len(real_sessions),
        "partial_blocks":   len(partial_blocks),
        "safety_net":       len(safety_net_blocks),
        "ps_core_blocks":   len(ps_core_blocks),
        "parse_errors":     len(errors),
    }


# ── ps_extraction.json ────────────────────────────────────────────────────────

def write_json_export(
    sessions: list,
    sb_index: dict,
    analysis_results: dict,
    summary: dict,
    output_path,
    source_files: list,
    errors: list,
) -> None:
    """Write ps_extraction.json for SIEM/tool ingestion."""
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    def _serialize_event(ev: dict) -> dict:
        """Flatten one session event to a compact JSON-friendly dict."""
        eid = ev.get("event_id")
        base = {
            "event_id":  eid,
            "timestamp": _fmt_ts(ev.get("timestamp", "")),
            "pid":       ev.get("pid", ""),
        }
        if eid == 400:
            base.update({
                "new_engine_state":  ev.get("new_engine_state", ""),
                "prev_engine_state": ev.get("prev_engine_state", ""),
                "host_application":  ev.get("host_application", ""),
            })
        elif eid == 403:
            base.update({
                "new_engine_state":  ev.get("new_engine_state", ""),
                "prev_engine_state": ev.get("prev_engine_state", ""),
            })
        elif eid == 600:
            base["provider_name"] = ev.get("provider_name", "")
        elif eid == 4103:
            base.update({
                "command_name": ev.get("command_name", ""),
                "command_type": ev.get("command_type", ""),
                "script_name":  ev.get("script_name", ""),
                "user":         ev.get("user", ""),
            })
        elif eid == 4104:
            sbid = ev.get("script_block_id", "")
            key = sbid if sbid else ""
            res = analysis_results.get(key)
            base.update({
                "script_block_id": sbid,
                "path":            ev.get("path", "") or "(in-memory)",
                "message_number":  ev.get("message_number", 1),
                "message_total":   ev.get("message_total", 1),
                "indicators":      res.indicator_flags() if res else [],
                "att_ck":          res.att_ck_techniques if res else ["T1059.001"],
            })
        elif eid == 800:
            base.update({
                "command_name": ev.get("command_name", ""),
                "command_line": ev.get("command_line", ""),
            })
        return base

    sessions_data = []
    for sess in sessions:
        if sess.host_id == "_unsessioned":
            continue
        if _is_ghost_session(sess):
            continue  # corrupt EID 400 with no metadata — omit from export
        sessions_data.append({
            "host_id":          sess.host_id,
            "pid":              sess.pid,
            "computer":         sess.computer,
            "start_ts":         _fmt_ts(sess.start_ts),
            "stop_ts":          _fmt_ts(sess.stop_ts),
            "duration":         sess.duration_str(),
            "host_name":        sess.host_name,
            "host_version":     sess.host_version,
            "host_application": sess.host_application,
            "encoded_command":  sess.encoded_command,
            "runspace_id":      sess.runspace_id,
            "user_sid":         sess.user_sid,
            "user_name":        sess.user_name,
            "providers":        sess.providers,
            "event_count":      len(sess.session_events),
            "events":           [_serialize_event(ev) for ev in sess.session_events],
        })

    blocks_data = {}
    for key, acc in sb_index.items():
        result = analysis_results.get(key)
        assembled = acc.assemble()
        blocks_data[key] = {
            "script_block_id":   acc.script_block_id,
            "path":              acc.path,
            "computer":          acc.computer,
            "first_seen":        _fmt_ts(acc.first_timestamp),
            "last_seen":         _fmt_ts(acc.last_timestamp),
            "fragment_count":    len(acc.fragments),
            "expected_total":    acc.expected_total,
            "is_complete":       acc.is_complete,
            "safety_net":        acc.was_safety_net_triggered,
            "assembled":         assembled,
            "indicators":        result.indicator_flags() if result else [],
            "att_ck_techniques": result.att_ck_techniques if result else ["T1059.001"],
            "detected_patterns": sorted(set(result.detected_patterns)) if result else [],
        }

    payload = {
        "metadata": {
            "tool":      "EventHawk",
            "version":   "ps_extractor/1.0",
            "timestamp": now,
            "sources":   source_files,
        },
        "summary":       summary,
        "sessions":      sessions_data,
        "script_blocks": blocks_data,
        "parse_errors":  errors[:200],
    }

    output_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False),
        encoding="utf-8",
        errors="replace",
    )


# ── ps_timeline.xlsx ──────────────────────────────────────────────────────────

def _make_safe_key(key: str) -> str:
    """Reproduce the same filename-safe key used by write_scriptblock_files."""
    safe = key.replace("{", "").replace("}", "").replace("/", "_").replace("\\", "_")
    if len(safe) > 200:
        safe = safe[:200]
    return safe


def write_csv_timeline(
    sessions: list,
    output_path,
    sb_index: dict | None = None,
) -> None:
    """Write ps_timeline.xlsx -- flat chronological event timeline.

    Output is a native Excel workbook (.xlsx) so that hyperlinks in the
    ``script_block_file`` column are activated directly by Excel — no formula
    injection warnings, no manual trust steps.  EID 4104 rows get a real
    clickable hyperlink that opens the corresponding ``scriptblock_<GUID>.txt``
    in the same output folder.  Short contextual detail for other EIDs goes
    into the ``detail`` column (capped at 500 chars).
    """
    import re
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    _EVENT_TYPE = {
        400:  "ENGINE_START",
        403:  "ENGINE_STOP",
        600:  "PROVIDER_START",
        800:  "PIPELINE_DETAIL",
        4103: "COMMAND_EXEC",
        4104: "SCRIPT_BLOCK",
    }

    _DETAIL_MAX = 500
    _CTRL_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')

    def _sanitize(v, max_len: int = 0) -> str:
        if v is None:
            return ""
        s = str(v)
        if not s:
            return ""
        s = _CTRL_RE.sub("", s)
        s = " ".join(s.split())
        if max_len and len(s) > max_len:
            s = s[:max_len] + "..."
        return s

    HEADERS = [
        "timestamp", "event_id", "event_type", "pid",
        "session_id", "user_sid", "user_name",
        "command", "script_block_id", "script_block_file", "detail",
    ]
    # Approximate column widths (characters)
    COL_WIDTHS = [22, 10, 16, 10, 38, 24, 20, 35, 38, 20, 55]

    wb = Workbook()
    ws = wb.active
    ws.title = "PS Timeline"

    # Header row styling
    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    ws.append(HEADERS)
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = hdr_align
        ws.column_dimensions[cell.column_letter].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 18

    # Hyperlink style for script_block_file column
    link_font = Font(color="0563C1", underline="single")
    # Column index of script_block_file (1-based)
    SB_FILE_COL = HEADERS.index("script_block_file") + 1

    _sb_index = sb_index or {}

    all_rows: list[tuple] = []
    for sess in sessions:
        s_id  = sess.host_id
        usid  = sess.user_sid
        uname = sess.user_name
        for ev in sess.session_events:
            eid     = ev.get("event_id", "")
            ts      = _fmt_ts(ev.get("timestamp", ""))
            pid     = ev.get("pid", "")
            etype   = _EVENT_TYPE.get(eid, str(eid))
            command = ""
            sbid    = ""
            sb_href = ""   # actual hyperlink target (relative filename)
            detail  = ""

            if eid == 4103:
                command = ev.get("command_name", "")
                detail  = ev.get("payload", "") or ev.get("user_data", "")
            elif eid == 800:
                command = ev.get("command_name", "")
                detail  = ev.get("command_line", "")
            elif eid == 4104:
                sbid    = ev.get("script_block_id", "")
                path    = ev.get("path", "")
                command = path if path else "(in-memory)"
                key = sbid
                if not key or key == "00000000-0000-0000-0000-000000000000":
                    rid = ev.get("event_record_id", 0)
                    key = f"{sbid}_{rid}"
                acc = _sb_index.get(key)
                if acc:
                    safe = _make_safe_key(key)
                    sb_href = f"scriptblock_{safe}.txt"
            elif eid == 600:
                command = ev.get("provider_name", "")
            elif eid == 400:
                detail = ev.get("host_application", "")

            all_rows.append((
                ts, eid, etype, pid, s_id, usid, uname,
                command, sbid, sb_href, detail,
            ))

    all_rows.sort(key=lambda r: r[0])

    for row_data in all_rows:
        row_num = ws.max_row + 1
        ws.append([
            _sanitize(row_data[0]),                        # timestamp
            _sanitize(row_data[1]),                        # event_id
            _sanitize(row_data[2]),                        # event_type
            _sanitize(row_data[3]),                        # pid
            _sanitize(row_data[4]),                        # session_id
            _sanitize(row_data[5]),                        # user_sid
            _sanitize(row_data[6]),                        # user_name
            _sanitize(row_data[7]),                        # command
            _sanitize(row_data[8]),                        # script_block_id
            "Open Script Block" if row_data[9] else "",   # script_block_file (text; hyperlink set below)
            _sanitize(row_data[10], max_len=_DETAIL_MAX), # detail
        ])
        if row_data[9]:  # sb_href present — attach real hyperlink
            cell = ws.cell(row=row_num, column=SB_FILE_COL)
            cell.hyperlink = row_data[9]
            cell.font = link_font

    wb.save(output_path)

